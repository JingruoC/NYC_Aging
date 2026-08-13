from __future__ import annotations

from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import select

from .db import Base, engine, SessionLocal
from .models import (
    HistoricalMenu,
    HistoricalMenuItem,
    HomeUpdate,
    Menu,
    MenuComment,
    MenuItem,
    NutrientThreshold,
    Recipe,
    RecipeAttachment,
    RecipeHomeCategorySetting,
)


HOME_UPDATES = [
    {
        "id": 1,
        "title": "Cycle menu preparation",
        "update_type": "Announcement",
        "summary": "Confirm cycle, service days, diet type, and contract selection before building a new menu.",
        "content": "Before starting a new cycle, confirm the contract, service days, diet type, and menu coverage.\n\nReview the cycle dates and selected meal types before recipes are placed. This helps prevent a completed week from being attached to the wrong contract or service pattern.\n\nIf an existing menu will be reused, open it from Menu List or Sample Menus and select it as the starting point before making changes.",
        "published_on": date.today() - timedelta(days=6),
        "image_source": "/images/cycle-menu-notice.svg",
    },
    {
        "id": 2,
        "title": "Seasonal produce planning",
        "update_type": "Story",
        "summary": "A short planning note on using approved seasonal fruit and vegetable recipes across the cycle.",
        "content": "Seasonal produce can add variety across a multi-week menu cycle while keeping the recipe catalog familiar to providers.\n\nUse the approved recipe list to compare fruit and vegetable choices, then check for repeated recipes across the selected week. Staff can feature approved seasonal recipes on the home page when they are especially useful for an upcoming cycle.\n\nRecipes still need to meet the applicable meal-component and nutrition requirements before the week can be marked complete.",
        "published_on": date.today() - timedelta(days=13),
        "image_source": "/images/seasonal-produce-story.svg",
    },
]

RECIPE_HOME_CATEGORIES = [
    ("breakfast", True),
    ("appetizer", False),
    ("entree", True),
    ("side", True),
    ("grain", True),
    ("vegetable", True),
    ("fruit", True),
    ("milk", True),
    ("dairy-free", False),
    ("condiments", False),
    ("juice", False),
    ("plant-based", True),
]


RECIPES = [
    {"recipe_id": 1, "recipe_name": "Cinnamon Oatmeal with Apples", "meal_type": "breakfast", "category": "grain", "calories": 180, "sodium_mg": 120, "protein_g": 6, "fiber_g": 5, "fat_g": 3, "tags": ["whole grain", "vegetarian", "low sodium"], "is_approved": True},
    {"recipe_id": 2, "recipe_name": "Scrambled Eggs with Spinach", "meal_type": "breakfast", "category": "entree", "calories": 210, "sodium_mg": 330, "protein_g": 14, "fiber_g": 2, "fat_g": 14, "tags": ["high protein"], "is_approved": True},
    {"recipe_id": 3, "recipe_name": "Whole Wheat Toast", "meal_type": "breakfast", "category": "grain", "calories": 110, "sodium_mg": 170, "protein_g": 4, "fiber_g": 3, "fat_g": 1, "tags": ["whole grain"], "is_approved": True},
    {"recipe_id": 4, "recipe_name": "Low Fat Milk", "meal_type": "breakfast", "category": "milk", "calories": 90, "sodium_mg": 130, "protein_g": 8, "fiber_g": 0, "fat_g": 2, "tags": ["dairy"], "is_approved": True},
    {"recipe_id": 5, "recipe_name": "Turkey Sausage Patty", "meal_type": "breakfast", "category": "entree", "calories": 150, "sodium_mg": 420, "protein_g": 11, "fiber_g": 0, "fat_g": 11, "tags": ["protein"], "is_approved": True},
    {"recipe_id": 6, "recipe_name": "Blueberry Yogurt Parfait", "meal_type": "breakfast", "category": "fruit", "calories": 160, "sodium_mg": 85, "protein_g": 9, "fiber_g": 3, "fat_g": 3, "tags": ["dairy", "fruit"], "is_approved": True},
    {"recipe_id": 7, "recipe_name": "Chicken and Brown Rice Bowl", "meal_type": "lunch", "category": "entree", "calories": 340, "sodium_mg": 520, "protein_g": 24, "fiber_g": 4, "fat_g": 10, "tags": ["high protein"], "is_approved": True},
    {"recipe_id": 8, "recipe_name": "Roasted Sweet Potatoes", "meal_type": "lunch", "category": "vegetable", "calories": 120, "sodium_mg": 65, "protein_g": 2, "fiber_g": 4, "fat_g": 4, "tags": ["vegetarian", "low sodium"], "is_approved": True},
    {"recipe_id": 9, "recipe_name": "Steamed Broccoli", "meal_type": "lunch", "category": "vegetable", "calories": 55, "sodium_mg": 40, "protein_g": 4, "fiber_g": 4, "fat_g": 1, "tags": ["low sodium"], "is_approved": True},
    {"recipe_id": 10, "recipe_name": "Brown Rice Pilaf", "meal_type": "lunch", "category": "grain", "calories": 160, "sodium_mg": 140, "protein_g": 4, "fiber_g": 3, "fat_g": 3, "tags": ["whole grain"], "is_approved": True},
    {"recipe_id": 11, "recipe_name": "Apple Slices", "meal_type": "lunch", "category": "fruit", "calories": 60, "sodium_mg": 0, "protein_g": 0, "fiber_g": 3, "fat_g": 0, "tags": ["fresh fruit", "low sodium"], "is_approved": True},
    {"recipe_id": 12, "recipe_name": "Garden Salad with Italian Dressing", "meal_type": "lunch", "category": "vegetable", "calories": 80, "sodium_mg": 120, "protein_g": 2, "fiber_g": 2, "fat_g": 5, "tags": ["vegetarian"], "is_approved": True},
    {"recipe_id": 13, "recipe_name": "Baked Fish with Lemon", "meal_type": "dinner", "category": "entree", "calories": 280, "sodium_mg": 260, "protein_g": 27, "fiber_g": 0, "fat_g": 12, "tags": ["high protein", "low sodium"], "is_approved": True},
    {"recipe_id": 14, "recipe_name": "Mashed Potatoes", "meal_type": "dinner", "category": "side", "calories": 150, "sodium_mg": 180, "protein_g": 3, "fiber_g": 2, "fat_g": 5, "tags": ["comfort food"], "is_approved": True},
    {"recipe_id": 15, "recipe_name": "Green Beans Almondine", "meal_type": "dinner", "category": "vegetable", "calories": 95, "sodium_mg": 55, "protein_g": 3, "fiber_g": 4, "fat_g": 5, "tags": ["low sodium"], "is_approved": True},
    {"recipe_id": 16, "recipe_name": "Whole Wheat Dinner Roll", "meal_type": "dinner", "category": "grain", "calories": 120, "sodium_mg": 150, "protein_g": 4, "fiber_g": 2, "fat_g": 2, "tags": ["whole grain"], "is_approved": True},
    {"recipe_id": 17, "recipe_name": "Fresh Orange", "meal_type": "dinner", "category": "fruit", "calories": 70, "sodium_mg": 0, "protein_g": 1, "fiber_g": 3, "fat_g": 0, "tags": ["fresh fruit"], "is_approved": True},
    {"recipe_id": 18, "recipe_name": "Baked Chicken Thigh", "meal_type": "dinner", "category": "entree", "calories": 310, "sodium_mg": 430, "protein_g": 28, "fiber_g": 0, "fat_g": 18, "tags": ["protein"], "is_approved": True},
    {"recipe_id": 19, "recipe_name": "Turkey Meatloaf", "meal_type": "dinner", "category": "entree", "calories": 290, "sodium_mg": 500, "protein_g": 26, "fiber_g": 1, "fat_g": 15, "tags": ["high protein"], "is_approved": True},
    {"recipe_id": 20, "recipe_name": "Carrot Raisin Salad", "meal_type": "lunch", "category": "vegetable", "calories": 110, "sodium_mg": 95, "protein_g": 1, "fiber_g": 3, "fat_g": 4, "tags": ["sweet", "vegetarian"], "is_approved": True},
    {"recipe_id": 21, "recipe_name": "Chicken Noodle Soup", "meal_type": "lunch", "category": "entree", "calories": 220, "sodium_mg": 670, "protein_g": 16, "fiber_g": 2, "fat_g": 7, "tags": ["warm meal"], "is_approved": True},
    {"recipe_id": 22, "recipe_name": "Low Fat Cottage Cheese", "meal_type": "snack", "category": "milk", "calories": 100, "sodium_mg": 360, "protein_g": 13, "fiber_g": 0, "fat_g": 2, "tags": ["high protein"], "is_approved": True},
    {"recipe_id": 23, "recipe_name": "Whole Grain Crackers", "meal_type": "snack", "category": "grain", "calories": 130, "sodium_mg": 150, "protein_g": 3, "fiber_g": 4, "fat_g": 3, "tags": ["whole grain"], "is_approved": True},
    {"recipe_id": 24, "recipe_name": "Peanut Butter Banana Sandwich", "meal_type": "snack", "category": "entree", "calories": 230, "sodium_mg": 210, "protein_g": 8, "fiber_g": 4, "fat_g": 10, "tags": ["protein", "fruit"], "is_approved": True},
    {"recipe_id": 25, "recipe_name": "Stewed Pears", "meal_type": "snack", "category": "fruit", "calories": 85, "sodium_mg": 15, "protein_g": 1, "fiber_g": 3, "fat_g": 0, "tags": ["low sodium"], "is_approved": True},
    {"recipe_id": 26, "recipe_name": "Vegetable Chili", "meal_type": "dinner", "category": "entree", "calories": 260, "sodium_mg": 410, "protein_g": 15, "fiber_g": 9, "fat_g": 7, "tags": ["high fiber", "vegetarian"], "is_approved": True},
    {"recipe_id": 27, "recipe_name": "Cornbread Muffin", "meal_type": "dinner", "category": "grain", "calories": 170, "sodium_mg": 240, "protein_g": 3, "fiber_g": 2, "fat_g": 5, "tags": ["whole grain"], "is_approved": True},
    {"recipe_id": 28, "recipe_name": "Cauliflower Mash", "meal_type": "dinner", "category": "vegetable", "calories": 75, "sodium_mg": 50, "protein_g": 3, "fiber_g": 3, "fat_g": 2, "tags": ["low sodium", "vegetarian"], "is_approved": True},
    {"recipe_id": 29, "recipe_name": "Baked Apples with Cinnamon", "meal_type": "snack", "category": "fruit", "calories": 95, "sodium_mg": 20, "protein_g": 1, "fiber_g": 4, "fat_g": 1, "tags": ["dessert"], "is_approved": True},
    {"recipe_id": 30, "recipe_name": "Cheddar Cheese Slice", "meal_type": "snack", "category": "milk", "calories": 110, "sodium_mg": 180, "protein_g": 7, "fiber_g": 0, "fat_g": 9, "tags": ["dairy"], "is_approved": True},
    {"recipe_id": 31, "recipe_name": "Lentil Stew", "meal_type": "lunch", "category": "entree", "calories": 240, "sodium_mg": 290, "protein_g": 16, "fiber_g": 10, "fat_g": 4, "tags": ["high fiber", "vegetarian"], "is_approved": True},
    {"recipe_id": 32, "recipe_name": "Spinach Salad", "meal_type": "lunch", "category": "vegetable", "calories": 70, "sodium_mg": 45, "protein_g": 2, "fiber_g": 3, "fat_g": 4, "tags": ["low sodium"], "is_approved": True},
    {"recipe_id": 33, "recipe_name": "Turkey Sandwich on Wheat", "meal_type": "lunch", "category": "entree", "calories": 320, "sodium_mg": 560, "protein_g": 19, "fiber_g": 4, "fat_g": 8, "tags": ["whole grain"], "is_approved": True},
    {"recipe_id": 34, "recipe_name": "Mixed Berry Cup", "meal_type": "lunch", "category": "fruit", "calories": 65, "sodium_mg": 0, "protein_g": 1, "fiber_g": 4, "fat_g": 0, "tags": ["fresh fruit"], "is_approved": True},
    {"recipe_id": 35, "recipe_name": "Rice Pudding", "meal_type": "snack", "category": "grain", "calories": 140, "sodium_mg": 110, "protein_g": 4, "fiber_g": 1, "fat_g": 4, "tags": ["dessert"], "is_approved": True},
    {"recipe_id": 36, "recipe_name": "Grilled Tofu with Sesame", "meal_type": "dinner", "category": "entree", "calories": 220, "sodium_mg": 340, "protein_g": 18, "fiber_g": 2, "fat_g": 11, "tags": ["vegetarian", "high protein"], "is_approved": True},
    {"recipe_id": 37, "recipe_name": "Barley Pilaf", "meal_type": "dinner", "category": "grain", "calories": 155, "sodium_mg": 95, "protein_g": 4, "fiber_g": 5, "fat_g": 2, "tags": ["whole grain", "low sodium"], "is_approved": True},
    {"recipe_id": 38, "recipe_name": "Honey Glazed Carrots", "meal_type": "dinner", "category": "vegetable", "calories": 85, "sodium_mg": 75, "protein_g": 1, "fiber_g": 3, "fat_g": 2, "tags": ["vegetarian"], "is_approved": True},
    {"recipe_id": 39, "recipe_name": "Fruit Salad Cup", "meal_type": "snack", "category": "fruit", "calories": 75, "sodium_mg": 10, "protein_g": 1, "fiber_g": 3, "fat_g": 0, "tags": ["fresh fruit", "low sodium"], "is_approved": True},
    {"recipe_id": 40, "recipe_name": "Egg Salad Sandwich", "meal_type": "lunch", "category": "entree", "calories": 310, "sodium_mg": 530, "protein_g": 14, "fiber_g": 3, "fat_g": 18, "tags": ["protein"], "is_approved": True},
    {"recipe_id": 44, "recipe_name": "Low Sodium Tomato Salsa", "meal_type": "lunch", "category": "condiments", "calories": 25, "sodium_mg": 75, "protein_g": 1, "fiber_g": 1, "fat_g": 0, "tags": ["condiments", "low sodium"], "is_approved": True},
    {"recipe_id": 45, "recipe_name": "Apple Juice Cup", "meal_type": "lunch", "category": "juice-dessert", "calories": 60, "sodium_mg": 5, "protein_g": 0, "fiber_g": 0, "fat_g": 0, "tags": ["juice"], "is_approved": True},
    {"recipe_id": 46, "recipe_name": "Vanilla Yogurt Cup", "meal_type": "breakfast", "category": "milk", "calories": 120, "sodium_mg": 95, "protein_g": 7, "fiber_g": 0, "fat_g": 2, "tags": ["dairy"], "is_approved": True},
    {"recipe_id": 47, "recipe_name": "Cranberry Juice Cup", "meal_type": "lunch", "category": "juice-dessert", "calories": 70, "sodium_mg": 10, "protein_g": 0, "fiber_g": 0, "fat_g": 0, "tags": ["juice"], "is_approved": True},
    {"recipe_id": 48, "recipe_name": "Low Sodium Italian Dressing", "meal_type": "lunch", "category": "condiments", "calories": 35, "sodium_mg": 90, "protein_g": 0, "fiber_g": 0, "fat_g": 3, "tags": ["condiments", "low sodium"], "is_approved": True},
    {"recipe_id": 49, "recipe_name": "Mustard Packet", "meal_type": "lunch", "category": "condiments", "calories": 5, "sodium_mg": 55, "protein_g": 0, "fiber_g": 0, "fat_g": 0, "tags": ["condiments"], "is_approved": True},
]


def _legacy_ingredients(row: dict) -> list[str]:
    name = row["recipe_name"].lower()
    category = row["category"]
    if "oatmeal" in name:
        return ["12 cups rolled oats", "10 pounds apples", "1/2 cup cinnamon", "3 gallons low fat milk", "2 cups brown sugar"]
    if "egg" in name:
        return ["50 eggs", "5 pounds spinach", "50 slices whole wheat bread", "3 cups light mayonnaise", "2 tablespoons black pepper"]
    if "milk" in name:
        return ["50 cups 1% low fat milk"]
    if "yogurt" in name:
        return ["25 pounds low fat yogurt", "8 pounds blueberries", "6 pounds whole grain granola"]
    if "rice" in name:
        return ["12 cups brown rice", "6 gallons low sodium vegetable stock", "5 pounds onion", "2 cups parsley"]
    if "tofu" in name:
        return ["18 pounds firm tofu", "2 cups sesame seeds", "3 cups low sodium soy sauce", "2 pounds scallions"]
    if "lentil" in name:
        return ["12 pounds lentils", "5 pounds carrots", "4 pounds celery", "8 pounds tomatoes", "5 gallons low sodium broth"]
    if "chicken" in name:
        return ["22 pounds chicken", "1 cup herbs", "1/2 cup garlic", "3 gallons low sodium broth"]
    if "turkey" in name:
        return ["22 pounds turkey", "5 pounds whole wheat bread crumbs", "4 pounds onion", "1 cup herbs"]
    if "fish" in name:
        return ["22 pounds white fish", "25 lemons", "2 cups parsley", "3 cups olive oil"]
    if category == "fruit":
        return [f"50 servings {row['recipe_name'].lower()}", "25 pounds fresh fruit"]
    if category == "vegetable":
        return [f"25 pounds {row['recipe_name'].lower()}", "20 pounds vegetables", "3 cups olive oil", "1 cup herbs"]
    if category == "grain":
        return [f"50 servings {row['recipe_name'].lower()}", "12 pounds whole grain ingredient", "1 cup low sodium seasoning"]
    if category == "milk":
        return [f"50 servings {row['recipe_name'].lower()}", "50 cups dairy or calcium-fortified equivalent"]
    return [f"50 servings {row['recipe_name'].lower()}", "20 pounds approved menu ingredient", "1 cup low sodium seasoning"]


def _legacy_claims(row: dict) -> list[str]:
    claims = set(row.get("tags", []))
    if row["sodium_mg"] <= 140:
        claims.add("low sodium")
    if row["protein_g"] >= 15:
        claims.add("high protein")
    if row["fiber_g"] >= 5:
        claims.add("high fiber")
    if row["category"] == "grain" and any("whole" in tag for tag in row.get("tags", [])):
        claims.add("whole grain")
    if row["category"] in {"fruit", "vegetable"}:
        claims.add("good source of vitamin c")
    if row["category"] == "milk":
        claims.add("good source of calcium")
    return sorted(claims)


def _vitamin_c(row: dict) -> float:
    if row["category"] == "fruit":
        return 35.0
    if row["category"] == "vegetable":
        return 28.0
    return 2.0


def _calcium(row: dict) -> float:
    if row["category"] == "milk":
        return 300.0
    if row["category"] == "vegetable":
        return 45.0
    return 15.0


for recipe in RECIPES:
    recipe.update(
        {
            "ingredients": _legacy_ingredients(recipe),
            "instructions": [
                "Review approved production record before service.",
                "Prepare according to approved Simple Servings recipe procedure.",
                "Hold and serve using the portion size listed for the menu.",
            ],
            "serving_size": "1 serving",
            "yield_servings": 50,
            "scale_note": "Use Scale to serving(s) in the recipe view to adjust production quantities.",
            "contributed_by": "NYC Aging Nutrition Unit",
            "created_on": date(2025, (recipe["recipe_id"] % 12) + 1, (recipe["recipe_id"] % 24) + 1),
            "is_public": True,
            "is_favorite": recipe["recipe_id"] in {1, 4, 7, 10, 11, 13, 31, 33, 34, 37},
            "is_dead": False,
            "nutrient_claims": _legacy_claims(recipe),
            "vitamin_c_mg": _vitamin_c(recipe),
            "calcium_mg": _calcium(recipe),
        }
    )

RECIPES.extend(
    [
        {
            "recipe_id": 41,
            "recipe_name": "Three Bean Vegetable Chili",
            "meal_type": "lunch",
            "category": "entree",
            "calories": 275,
            "sodium_mg": 410,
            "protein_g": 14,
            "fiber_g": 11,
            "fat_g": 6,
            "tags": ["pending review", "Full recipe with ingredients and directions", "plant-based", "high fiber", "entree"],
            "is_approved": False,
            "ingredients": ["low sodium kidney beans", "black beans", "pinto beans", "tomatoes", "onions", "bell peppers", "chili powder"],
            "instructions": ["Combine vegetables, beans, tomatoes, and seasoning.", "Simmer until vegetables are tender.", "Hold hot and serve one measured portion."],
            "serving_size": "1 cup",
            "yield_servings": 60,
            "scale_note": "\n".join(
                [
                    "Submission type: Full recipe with ingredients and directions",
                    "Contact Person: Astoria NSC",
                    "Preferred E-mail Address: menu.staff@astoriansc.example",
                    "Can be made public: Yes",
                    "Recipe source: A recipe created by our organization",
                    "Recipe file names: three-bean-vegetable-chili.pdf",
                    "Comments: Provider requests review for plant-based entree rotation.",
                ]
            ),
            "contributed_by": "Astoria NSC",
            "created_on": date(2026, 7, 18),
            "is_public": True,
            "is_favorite": False,
            "is_dead": False,
            "nutrient_claims": ["high fiber", "plant-based entree"],
            "vitamin_c_mg": 22,
            "calcium_mg": 75,
        },
        {
            "recipe_id": 42,
            "recipe_name": "Southwest Brown Rice Bowl",
            "meal_type": "lunch",
            "category": "grain",
            "calories": 365,
            "sodium_mg": 690,
            "protein_g": 12,
            "fiber_g": 7,
            "fat_g": 9,
            "tags": ["pending review", "Recipe nutrition analysis", "whole grain", "grain"],
            "is_approved": False,
            "ingredients": ["Vegetables per serving: corn 1/4 cup, peppers 1/4 cup", "Plant-based protein per serving: black beans 1/4 cup", "Fruit per serving: 0"],
            "instructions": ["Review the submitted recipe nutrition analysis panel."],
            "serving_size": "1 serving from analysis",
            "yield_servings": 48,
            "scale_note": "\n".join(
                [
                    "Submission type: Recipe nutrition analysis",
                    "Contact Person: ABSW OAC",
                    "Preferred E-mail Address: nutrition@abswoac.example",
                    "Can be made public: Yes",
                    "Recipe Yield: 48 servings",
                    "Number of Servings: 48",
                    "Nutrition Facts Panel file names: southwest-brown-rice-analysis.xlsx",
                    "Vegetables Per Serving: corn 1/4 cup, peppers 1/4 cup",
                    "Plant-Based Protein Per Serving: black beans 1/4 cup",
                    "Fruit Per Serving: 0",
                ]
            ),
            "contributed_by": "ABSW OAC",
            "created_on": date(2026, 7, 20),
            "is_public": True,
            "is_favorite": False,
            "is_dead": False,
            "nutrient_claims": ["whole grain", "good source of fiber"],
            "vitamin_c_mg": 18,
            "calcium_mg": 40,
        },
        {
            "recipe_id": 43,
            "recipe_name": "Reduced Sodium Turkey Patty",
            "meal_type": "breakfast",
            "category": "entree",
            "calories": 145,
            "sodium_mg": 360,
            "protein_g": 13,
            "fiber_g": 0,
            "fat_g": 8,
            "tags": ["pending review", "Pre-prepared product label", "protein", "breakfast"],
            "is_approved": False,
            "ingredients": ["Manufacturer: Metro Food Service", "Serving size: 1 patty", "Nutrition Facts panel files: reduced-sodium-turkey-patty-label.pdf"],
            "instructions": ["Review the pre-prepared product label and ingredient list."],
            "serving_size": "1 patty",
            "yield_servings": 1,
            "scale_note": "\n".join(
                [
                    "Submission type: Pre-prepared product label",
                    "Contact Person: Ridgewood OAC",
                    "Preferred E-mail Address: catering@ridgewoodoac.example",
                    "Can be made public: No",
                    "Manufacturer: Metro Food Service",
                    "Serving Size: 1 patty",
                    "Saturated Fat (g): 2.5",
                    "Trans Fat (g): 0",
                    "Cholesterol (mg): 35",
                    "Total Carbohydrate (g): 4",
                    "Total Sugars (g): 0",
                    "Added Sugars (g): 0",
                    "Vitamin D (mcg): 0",
                    "Iron (mg): 1.2",
                    "Potassium (mg): 180",
                    "Nutrition Facts Panel and Ingredient List file names: reduced-sodium-turkey-patty-label.pdf",
                ]
            ),
            "contributed_by": "Ridgewood OAC",
            "created_on": date(2026, 7, 21),
            "is_public": False,
            "is_favorite": False,
            "is_dead": False,
            "nutrient_claims": ["high protein"],
            "vitamin_c_mg": 0,
            "calcium_mg": 20,
        },
    ]
)


def _pdf_text(value: str) -> str:
    return value.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _build_mock_pdf(title: str, lines: list[str]) -> bytes:
    """Build a small valid PDF so seeded upload links can be tested end to end."""
    commands = ["BT", "/F1 18 Tf", f"72 748 Td ({_pdf_text(title)}) Tj", "/F1 11 Tf"]
    for line in lines:
        commands.extend(["0 -24 Td", f"({_pdf_text(line)}) Tj"])
    commands.append("ET")
    stream = "\n".join(commands).encode("latin-1")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
    ]
    document = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = [0]
    for index, obj in enumerate(objects, start=1):
        offsets.append(len(document))
        document.extend(f"{index} 0 obj\n".encode("ascii"))
        document.extend(obj)
        document.extend(b"\nendobj\n")
    xref_offset = len(document)
    document.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    document.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        document.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    document.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    return bytes(document)


def _build_mock_nutritionist_pro_workbook() -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Recipe Summary"
    summary.append(["Nutritionist Pro Recipe Analysis"])
    summary.append(["Recipe Name", "Southwest Brown Rice Bowl"])
    summary.append(["Recipe Yield", 48])
    summary.append(["Serving Size", "1 serving"])
    summary.append(["Contributor", "ABSW OAC"])
    summary.append(["Analysis Date", "2026-07-20"])

    nutrients = workbook.create_sheet("Nutrient Analysis")
    nutrients.append(["Nutrient", "Amount per serving", "Unit"])
    nutrient_rows = [
        ("Calories", 365, "kcal"),
        ("Total Fat", 9, "g"),
        ("Saturated Fat", 1.5, "g"),
        ("Trans Fat", 0, "g"),
        ("Cholesterol", 0, "mg"),
        ("Sodium", 690, "mg"),
        ("Total Carbohydrate", 62, "g"),
        ("Dietary Fiber", 7, "g"),
        ("Total Sugars", 5, "g"),
        ("Added Sugars", 0, "g"),
        ("Protein", 12, "g"),
        ("Vitamin D", 0, "mcg"),
        ("Calcium", 40, "mg"),
        ("Iron", 2.8, "mg"),
        ("Potassium", 510, "mg"),
    ]
    for row in nutrient_rows:
        nutrients.append(row)

    ingredients = workbook.create_sheet("Ingredients")
    ingredients.append(["Ingredient", "Quantity", "Unit"])
    ingredients.append(["Brown rice, cooked", 24, "cups"])
    ingredients.append(["Black beans, low sodium", 12, "cups"])
    ingredients.append(["Corn", 12, "cups"])
    ingredients.append(["Bell peppers", 12, "cups"])
    ingredients.append(["Salsa, low sodium", 6, "cups"])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            width = max(len(str(cell.value or "")) for cell in column) + 2
            sheet.column_dimensions[column[0].column_letter].width = min(width, 42)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _seed_recipe_attachments(session) -> None:
    attachment_rows = [
        {
            "recipe_id": 41,
            "file_name": "three-bean-vegetable-chili.pdf",
            "content_type": "application/pdf",
            "file_kind": "recipe_file",
            "uploaded_at": datetime(2026, 7, 18, 10, 15),
            "content": _build_mock_pdf(
                "Three Bean Vegetable Chili",
                [
                    "Submitted by Astoria NSC",
                    "Yield: 60 servings | Serving size: 1 cup",
                    "Ingredients and production directions attached for NYC Aging review.",
                ],
            ),
        },
        {
            "recipe_id": 42,
            "file_name": "southwest-brown-rice-analysis.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "file_kind": "nutrition_analysis",
            "uploaded_at": datetime(2026, 7, 20, 14, 30),
            "content": _build_mock_nutritionist_pro_workbook(),
        },
        {
            "recipe_id": 43,
            "file_name": "reduced-sodium-turkey-patty-label.pdf",
            "content_type": "application/pdf",
            "file_kind": "product_label",
            "uploaded_at": datetime(2026, 7, 21, 9, 45),
            "content": _build_mock_pdf(
                "Reduced Sodium Turkey Patty",
                [
                    "Manufacturer: Metro Food Service",
                    "Serving size: 1 patty | Calories: 145 | Sodium: 360 mg",
                    "Submitted by Ridgewood OAC as a pre-prepared product label.",
                ],
            ),
        },
    ]
    for row in attachment_rows:
        existing = session.scalar(
            select(RecipeAttachment).where(
                RecipeAttachment.recipe_id == row["recipe_id"],
                RecipeAttachment.file_name == row["file_name"],
            )
        )
        if existing is not None:
            continue
        content = row["content"]
        session.add(RecipeAttachment(**row, file_size=len(content)))


THRESHOLDS = [
    {"nutrient_key": "calories", "low_fail": 550, "low_warn": 650, "high_warn": 850, "high_fail": 950, "unit": "kcal"},
    {"nutrient_key": "sodium_mg", "low_fail": None, "low_warn": None, "high_warn": 1800, "high_fail": 2300, "unit": "mg"},
    {"nutrient_key": "protein_g", "low_fail": 20, "low_warn": 28, "high_warn": None, "high_fail": None, "unit": "g"},
    {"nutrient_key": "fiber_g", "low_fail": 7, "low_warn": 10, "high_warn": None, "high_fail": None, "unit": "g"},
    {"nutrient_key": "fat_g", "low_fail": None, "low_warn": None, "high_warn": 32, "high_fail": 40, "unit": "g"},
]


HISTORICAL_MENUS = [
    {"name": "Monday Breakfast A", "service_date": date(2026, 3, 2), "passes_nutrition": True, "items": [1, 3, 4, 6]},
    {"name": "Monday Lunch A", "service_date": date(2026, 3, 2), "passes_nutrition": True, "items": [7, 10, 9, 11]},
    {"name": "Tuesday Lunch A", "service_date": date(2026, 3, 3), "passes_nutrition": True, "items": [31, 37, 38, 34]},
    {"name": "Tuesday Dinner A", "service_date": date(2026, 3, 3), "passes_nutrition": True, "items": [13, 16, 15, 17]},
    {"name": "Wednesday Lunch A", "service_date": date(2026, 3, 4), "passes_nutrition": False, "items": [21, 10, 12, 30]},
    {"name": "Wednesday Snack A", "service_date": date(2026, 3, 4), "passes_nutrition": True, "items": [25, 23, 4]},
    {"name": "Thursday Breakfast A", "service_date": date(2026, 3, 5), "passes_nutrition": True, "items": [2, 3, 4, 11]},
    {"name": "Thursday Lunch A", "service_date": date(2026, 3, 5), "passes_nutrition": True, "items": [33, 32, 34]},
    {"name": "Friday Dinner A", "service_date": date(2026, 3, 6), "passes_nutrition": True, "items": [18, 37, 28, 17]},
    {"name": "Friday Lunch A", "service_date": date(2026, 3, 6), "passes_nutrition": False, "items": [40, 14, 30, 35]},
    {"name": "Monday Dinner B", "service_date": date(2026, 3, 9), "passes_nutrition": True, "items": [26, 37, 38, 39]},
    {"name": "Tuesday Breakfast B", "service_date": date(2026, 3, 10), "passes_nutrition": True, "items": [1, 6, 4]},
    {"name": "Tuesday Lunch B", "service_date": date(2026, 3, 10), "passes_nutrition": True, "items": [7, 8, 9, 11]},
    {"name": "Wednesday Dinner B", "service_date": date(2026, 3, 11), "passes_nutrition": False, "items": [19, 14, 27, 29]},
    {"name": "Thursday Lunch B", "service_date": date(2026, 3, 12), "passes_nutrition": True, "items": [31, 10, 32, 34]},
    {"name": "Thursday Snack B", "service_date": date(2026, 3, 12), "passes_nutrition": True, "items": [39, 23, 4]},
    {"name": "Friday Breakfast B", "service_date": date(2026, 3, 13), "passes_nutrition": True, "items": [2, 3, 6, 4]},
    {"name": "Friday Lunch B", "service_date": date(2026, 3, 13), "passes_nutrition": True, "items": [33, 12, 17]},
    {"name": "Monday Lunch C", "service_date": date(2026, 3, 16), "passes_nutrition": True, "items": [7, 10, 15, 11]},
    {"name": "Tuesday Dinner C", "service_date": date(2026, 3, 17), "passes_nutrition": True, "items": [36, 37, 28, 25]},
]

LEGACY_SAMPLE_NAMES = [
    "Sample 5-day Lunch Menu- Chinese 1",
    "Sample Breakfast- Cold",
    "Sample 7-day Lunch Menu- Polish 2",
    "Sample 7-day Lunch Menu- Russian 2",
    "Sample 7-day Lunch Menu- Vegetarian 1",
    "Sample 7-day Lunch Menu- Chinese 2",
    "Sample 7-day Lunch Menu- Kosher 1",
    "Sample 7-day Lunch Menu- Halal 2",
    "Sample 7-day Lunch Menu- Latin 1",
    "Sample 7-day Lunch Menu- Standard 1",
    "Sample 7-day Lunch Menu- Korean 1",
    "Sample 7-day Lunch Menu- Caribbean 1",
    "Sample 5-day Lunch Menu- Latin 2, Good For Chilled",
    "Sample 5-day Lunch Menu- Mediterranean 2",
    "Sample 5-day Lunch Menu- Cold",
    "Sample 5-day Lunch Menu- Standard 3",
    "NYC Aging Breakfast Hot Fiscal Year 2027",
    "NYC Aging Lunch Hot Fiscal Year 2027",
    "Gotham Catering Lunch Frozen Kosher Fiscal Year 2027",
    "Gotham Catering Lunch Frozen Vegetarian Fiscal Year 2027",
]

LEGACY_CONTRACTS = [
    "ABSW OAC",
    "Aging Through Arts Center (Encore at St Malachy's)",
    "Agudath Israel Brookdale Senior Center",
    "Agudath Israel Moriah Older Adult Luncheon Club",
    "Albany OAC",
    "Allen AME Community Senior Citizens Centers - Linden Blvd",
    "Allen AME Rockaway Blvd Senior Center",
    "ARC XVI Central Harlem Center",
    "Astoria NSC",
    "QUEENS COMM HOUSE HOME DELIVERED MEALS AT FOREST HILLS - QUEENS 3",
]


for index, menu in enumerate(HISTORICAL_MENUS, start=1):
    if index <= len(LEGACY_SAMPLE_NAMES):
        menu["name"] = LEGACY_SAMPLE_NAMES[index - 1]
    meal_name = menu["name"].lower()
    if "breakfast" in meal_name:
        meal_type = "Breakfast"
    elif "dinner" in meal_name:
        meal_type = "Dinner"
    elif "snack" in meal_name:
        meal_type = "Snack"
    else:
        meal_type = "Lunch"
    menu.update(
        {
            "program_type": "Home Delivered Meal" if "home delivered" in meal_name or "frozen" in meal_name else "Congregate",
            "meal_type": meal_type,
            "menu_coverage": f"{meal_type} only" if meal_type in {"Breakfast", "Lunch", "Dinner"} else "Lunch only",
            "diet_type": "Regular",
            "menu_duration_type": "Medically Tailored Meal - Very Low Sodium" if "very low sodium" in meal_name else "Regular",
            "meal_served_format": "Cold" if "cold" in meal_name else ("Frozen" if "frozen" in meal_name else "Hot"),
            "menu_tags": ["Vegetarian"] if "vegetarian" in meal_name else [],
            "cycle": "Spring/Summer" if index % 2 else "Fiscal Year",
            "days_per_week": 7 if "7-day" in meal_name else 5,
            "contracts": [LEGACY_CONTRACTS[(index - 1) % len(LEGACY_CONTRACTS)]],
            "sample_category": "Sample Menu" if index <= 12 else "Historical Menu",
        }
    )


SAMPLE_COMPONENT_RECIPES = {
    "Breakfast": {
        "entree": [2, 5, 2, 5, 2, 5, 2],
        "grains": [1, 3, 1, 3, 1, 3, 1],
        "vegetable": [2, 2, 2, 2, 2, 2, 2],
        "fruit": [6, 11, 6, 34, 6, 39, 6],
        "dairy": [4, 46, 4, 46, 4, 46, 4],
    },
    "Lunch": {
        "entree": [7, 31, 33, 40, 21, 7, 31],
        "grains": [10, 23, 10, 23, 10, 23, 10],
        "vegetable": [9, 8, 12, 20, 32, 9, 8],
        "fruit": [11, 34, 25, 39, 11, 34, 25],
        "dairy": [22, 30, 22, 30, 22, 30, 22],
    },
    "Dinner": {
        "entree": [13, 18, 19, 26, 36, 13, 18],
        "grains": [16, 37, 27, 16, 37, 27, 16],
        "vegetable": [15, 28, 38, 15, 28, 38, 15],
        "fruit": [17, 29, 39, 17, 29, 39, 17],
        "dairy": [22, 30, 22, 30, 22, 30, 22],
    },
}


def _complete_sample_items(
    meal_type: str,
    days_per_week: int,
    *,
    vegetarian: bool = False,
) -> list[dict[str, int | str]]:
    pools = dict(SAMPLE_COMPONENT_RECIPES.get(meal_type, SAMPLE_COMPONENT_RECIPES["Lunch"]))
    if vegetarian:
        pools["entree"] = {
            "Breakfast": [2, 2, 2, 2, 2, 2, 2],
            "Lunch": [31, 40, 31, 40, 31, 40, 31],
            "Dinner": [36, 26, 36, 26, 36, 26, 36],
        }.get(meal_type, [31, 40, 31, 40, 31, 40, 31])
    items: list[dict[str, int | str]] = []
    for day_index in range(days_per_week):
        for component_key, recipes in pools.items():
            items.append(
                {
                    "recipe_id": recipes[day_index % len(recipes)],
                    "day_index": day_index,
                    "component_key": component_key,
                }
            )
    return items


for menu in HISTORICAL_MENUS:
    sample_meal_type = menu["meal_type"] if menu["meal_type"] in SAMPLE_COMPONENT_RECIPES else "Lunch"
    menu["items"] = _complete_sample_items(
        sample_meal_type,
        menu["days_per_week"],
        vegetarian="vegetarian" in menu["name"].lower(),
    )


LEGACY_RECIPE_FIELDS = [
    "ingredients",
    "instructions",
    "serving_size",
    "yield_servings",
    "scale_note",
    "contributed_by",
    "created_on",
    "is_public",
    "is_favorite",
    "is_dead",
    "nutrient_claims",
    "vitamin_c_mg",
    "calcium_mg",
]


SAVED_MENU_COMPONENT_ORDER = [
    "appetizer",
    "entree",
    "grains",
    "vegetable",
    "fruit",
    "dairy",
    "juice-dessert",
    "condiments",
]


def _weekly_component_plan(days: list[dict[str, int | None]]) -> list[dict[str, int | str]]:
    placements: list[dict[str, int | str]] = []
    for day_index, day in enumerate(days):
        for component_key in SAVED_MENU_COMPONENT_ORDER:
            recipe_id = day.get(component_key)
            if recipe_id is None:
                continue
            placements.append(
                {
                    "recipe_id": recipe_id,
                    "day_index": day_index,
                    "component_key": component_key,
                }
            )
    return placements


SAVED_MENUS = [
    {
        "name": "ABSW OAC July Lunch Cycle - Week 1",
        "contract_name": "ABSW OAC",
        "program_type": "Congregate",
        "meal_type": "Lunch",
        "menu_coverage": "Lunch only",
        "diet_type": "Regular",
        "menu_format": "Weekly",
        "menu_duration_type": "Regular",
        "meal_served_format": "Hot",
        "menu_tags": [],
        "cycle": "Fiscal Year",
        "cycle_start_date": date(2026, 7, 1),
        "cycle_end_date": date(2027, 6, 30),
        "contracts": ["ABSW OAC"],
        "completed_weeks": [1],
        "submitted_programs": ["Congregate"],
        "status": "Submitted To NYC Aging",
        "status_date": date(2026, 7, 9),
        "submitted_to": "NYC Aging Nutrition Unit",
        "submitted_to_nyc_aging_on": date(2026, 7, 9),
        "nutrition_advisor": "NYC Aging Nutrition Unit",
        "created_by": "ABSW OAC",
        "service_date": date(2026, 7, 6),
        "start_date": date(2026, 7, 6),
        "end_date": date(2026, 7, 10),
        "days_per_week": 5,
        "cycle_week": 1,
        "notes": "Provider submitted Week 1 lunch cycle for review.",
        "returned_comments": None,
        "approval_notes": None,
        "is_favorite": True,
        "placements": _weekly_component_plan(
            [
                {"appetizer": 12, "entree": 7, "grains": 10, "vegetable": 9, "fruit": 11, "dairy": 4, "juice-dessert": 45, "condiments": 48},
                {"appetizer": 32, "entree": 31, "grains": 37, "vegetable": 8, "fruit": 34, "dairy": 46, "juice-dessert": 47, "condiments": 44},
                {"appetizer": 20, "entree": 33, "grains": 10, "vegetable": 12, "fruit": 17, "dairy": 4, "juice-dessert": 35, "condiments": 49},
                {"appetizer": 9, "entree": 40, "grains": 23, "vegetable": 20, "fruit": 39, "dairy": 22, "juice-dessert": 45, "condiments": 48},
                {"appetizer": 32, "entree": 21, "grains": 10, "vegetable": 8, "fruit": 25, "dairy": 30, "juice-dessert": 29, "condiments": 44},
            ]
        ),
        "comments": [
            {
                "created_at": datetime(2026, 7, 9, 10, 15),
                "action": "Submitted",
                "author": "ABSW OAC",
                "role": "Provider / Caterer",
                "body": "Submitted Week 1 lunch cycle with hot service format for NYC Aging review.",
                "visibility": "Admin and provider",
                "badge_class": "brand",
                "target_type": "menu",
                "target_label": "General menu summary",
                "review_status": "open",
            }
        ],
    },
    {
        "name": "ABSW OAC July Lunch Cycle - Week 2",
        "contract_name": "ABSW OAC",
        "program_type": "Congregate",
        "meal_type": "Lunch",
        "menu_coverage": "Lunch only",
        "diet_type": "Regular",
        "menu_format": "Weekly",
        "menu_duration_type": "Regular",
        "meal_served_format": "Hot",
        "menu_tags": [],
        "cycle": "Fiscal Year",
        "cycle_start_date": date(2026, 7, 1),
        "cycle_end_date": date(2027, 6, 30),
        "contracts": ["ABSW OAC"],
        "completed_weeks": [1, 2],
        "submitted_programs": ["Congregate"],
        "status": "Returned for correction (from NYC Aging)",
        "status_date": date(2026, 7, 11),
        "submitted_to": "ABSW OAC",
        "submitted_to_nyc_aging_on": date(2026, 7, 8),
        "nutrition_advisor": "Malek, Esther",
        "created_by": "ABSW OAC",
        "service_date": date(2026, 7, 13),
        "start_date": date(2026, 7, 13),
        "end_date": date(2026, 7, 17),
        "days_per_week": 5,
        "cycle_week": 2,
        "notes": "Returned to provider with sodium and component comments.",
        "returned_comments": "Please replace the high-sodium soup on Wednesday and confirm dairy or non-dairy equivalent for Friday.",
        "approval_notes": None,
        "is_favorite": False,
        "placements": _weekly_component_plan(
            [
                {"appetizer": 12, "entree": 21, "grains": 10, "vegetable": 9, "fruit": 11, "dairy": 4, "juice-dessert": 45, "condiments": 48},
                {"appetizer": 20, "entree": 7, "grains": 10, "vegetable": 8, "fruit": 34, "dairy": 46, "juice-dessert": 47, "condiments": 44},
                {"appetizer": 32, "entree": 40, "grains": 23, "vegetable": 12, "fruit": 25, "dairy": 30, "juice-dessert": 35, "condiments": 49},
                {"appetizer": 9, "entree": 31, "grains": 10, "vegetable": 20, "fruit": 39, "dairy": 4, "juice-dessert": 45, "condiments": 48},
                {"appetizer": 12, "entree": 33, "grains": 37, "vegetable": 15, "fruit": 17, "dairy": 22, "juice-dessert": 29, "condiments": 44},
            ]
        ),
        "comments": [
            {
                "created_at": datetime(2026, 7, 11, 9, 30),
                "action": "Returned correction",
                "author": "Malek, Esther",
                "role": "Admin / Department staff",
                "body": "Please replace the high-sodium soup on Wednesday and confirm dairy or non-dairy equivalent for Friday.",
                "visibility": "Admin and provider",
                "badge_class": "warn",
                "target_type": "component",
                "target_label": "Wednesday Lunch - Entrée",
                "day_index": 2,
                "meal_slot": "Lunch",
                "component_key": "entree",
                "review_status": "open",
            },
            {
                "created_at": datetime(2026, 7, 11, 9, 45),
                "action": "Internal note",
                "author": "NYC Aging Nutrition Unit",
                "role": "Admin / Department staff",
                "body": "Review sodium again after provider substitution. Wednesday lunch was the main concern.",
                "visibility": "Internal",
                "badge_class": "brand",
                "target_type": "nutrition",
                "target_label": "Sodium review",
                "nutrient_key": "sodium_mg",
                "review_status": "open",
            },
        ],
    },
    {
        "name": "ABSW OAC July Breakfast Cycle - Week 1",
        "contract_name": "ABSW OAC",
        "program_type": "Congregate",
        "meal_type": "Breakfast",
        "menu_coverage": "Breakfast only",
        "diet_type": "Regular",
        "menu_format": "Weekly",
        "menu_duration_type": "Regular",
        "meal_served_format": "Hot",
        "menu_tags": [],
        "cycle": "Fiscal Year",
        "cycle_start_date": date(2026, 7, 1),
        "cycle_end_date": date(2027, 6, 30),
        "contracts": ["ABSW OAC"],
        "completed_weeks": [1],
        "submitted_programs": ["Congregate"],
        "status": "Approved",
        "status_date": date(2026, 7, 10),
        "submitted_to": "ABSW OAC",
        "submitted_to_nyc_aging_on": date(2026, 7, 7),
        "nutrition_advisor": "NYC Aging Nutrition Unit",
        "created_by": "ABSW OAC",
        "service_date": date(2026, 7, 6),
        "start_date": date(2026, 7, 6),
        "end_date": date(2026, 7, 10),
        "days_per_week": 5,
        "cycle_week": 1,
        "notes": "Approved breakfast cycle.",
        "returned_comments": None,
        "approval_notes": "Approved for Week 1 breakfast service. Keep milk equivalent visible on printed menu.",
        "is_favorite": False,
        "placements": _weekly_component_plan(
            [
                {"entree": 1, "grains": 3, "fruit": 6, "dairy": 4, "juice-dessert": 45, "condiments": 48},
                {"entree": 2, "grains": 3, "vegetable": 9, "fruit": 11, "dairy": 46, "juice-dessert": 47, "condiments": 49},
                {"entree": 5, "grains": 1, "fruit": 6, "dairy": 4, "juice-dessert": 45, "condiments": 44},
                {"entree": 2, "grains": 3, "vegetable": 32, "fruit": 34, "dairy": 46, "juice-dessert": 47, "condiments": 48},
                {"entree": 1, "grains": 3, "fruit": 39, "dairy": 4, "juice-dessert": 45, "condiments": 49},
            ]
        ),
        "comments": [
            {
                "created_at": datetime(2026, 7, 10, 14, 0),
                "action": "Approval note",
                "author": "NYC Aging Nutrition Unit",
                "role": "Admin / Department staff",
                "body": "Approved for Week 1 breakfast service. Keep milk equivalent visible on printed menu.",
                "visibility": "Admin and provider",
                "badge_class": "good",
                "target_type": "menu",
                "target_label": "General menu summary",
                "review_status": "resolved",
            }
        ],
    },
    {
        "name": "Astoria NSC August Lunch Cycle - Week 1",
        "contract_name": "Astoria NSC",
        "program_type": "Congregate",
        "meal_type": "Lunch",
        "menu_coverage": "Lunch only",
        "diet_type": "Regular",
        "menu_format": "Weekly",
        "menu_duration_type": "Regular",
        "meal_served_format": "Fresh Chilled",
        "menu_tags": [],
        "cycle": "Summer",
        "cycle_start_date": date(2026, 8, 1),
        "cycle_end_date": date(2026, 8, 31),
        "contracts": ["Astoria NSC"],
        "completed_weeks": [1],
        "submitted_programs": ["Congregate"],
        "status": "Contract(s) Reviewed Menu",
        "status_date": date(2026, 7, 12),
        "submitted_to": "NYC Aging Nutrition Unit",
        "submitted_to_nyc_aging_on": date(2026, 7, 12),
        "nutrition_advisor": "Regional Nutrition Advisor",
        "created_by": "Astoria NSC",
        "service_date": date(2026, 8, 3),
        "start_date": date(2026, 8, 3),
        "end_date": date(2026, 8, 7),
        "days_per_week": 5,
        "cycle_week": 1,
        "notes": "Contract reviewed and ready for NYC Aging review.",
        "returned_comments": None,
        "approval_notes": None,
        "is_favorite": False,
        "placements": _weekly_component_plan(
            [
                {"appetizer": 32, "entree": 31, "grains": 10, "vegetable": 9, "fruit": 34, "dairy": 4, "juice-dessert": 45, "condiments": 48},
                {"appetizer": 12, "entree": 7, "grains": 10, "vegetable": 8, "fruit": 11, "dairy": 46, "juice-dessert": 47, "condiments": 44},
                {"appetizer": 20, "entree": 33, "grains": 23, "vegetable": 12, "fruit": 17, "dairy": 30, "juice-dessert": 35, "condiments": 49},
                {"appetizer": 9, "entree": 21, "grains": 10, "vegetable": 20, "fruit": 39, "dairy": 22, "juice-dessert": 45, "condiments": 48},
                {"appetizer": 32, "entree": 36, "grains": 37, "vegetable": 28, "fruit": 25, "dairy": 4, "juice-dessert": 29, "condiments": 44},
            ]
        ),
        "comments": [
            {
                "created_at": datetime(2026, 7, 12, 11, 20),
                "action": "Submitted",
                "author": "Astoria NSC",
                "role": "Provider / Caterer",
                "body": "Contract reviewed menu and submitted for NYC Aging review.",
                "visibility": "Admin and provider",
                "badge_class": "brand",
                "target_type": "menu",
                "target_label": "General menu summary",
                "review_status": "open",
            }
        ],
    },
]


def _component_for_recipe(recipe: Recipe | None, position: int) -> str:
    if recipe is None:
        return "entree" if position % 4 == 1 else "vegetable"
    return {
        "grain": "grains",
        "side": "vegetable",
        "milk": "dairy",
        "snack": "alternate",
    }.get(recipe.category, recipe.category)


def _slot_for_menu(menu: dict, component_key: str) -> str:
    meal_type = str(menu["meal_type"]).lower()
    return f"{meal_type}_{component_key}"


def _placements_for_saved_menu(menu: dict) -> list[dict[str, int | str]]:
    if "placements" in menu:
        return menu["placements"]

    placements: list[dict[str, int | str]] = []
    for position, recipe_id in enumerate(menu.get("items", []), start=1):
        placements.append(
            {
                "recipe_id": recipe_id,
                "day_index": (position - 1) // 4,
                "component_key": "",
            }
        )
    return placements


def _replace_saved_menu_items(session, menu: Menu, menu_row: dict) -> None:
    existing_items = session.scalars(select(MenuItem).where(MenuItem.menu_id == menu.id)).all()
    for item in existing_items:
        session.delete(item)
    session.flush()

    for position, placement in enumerate(_placements_for_saved_menu(menu_row), start=1):
        recipe = session.get(Recipe, placement["recipe_id"])
        component_key = placement.get("component_key") or _component_for_recipe(recipe, position)
        session.add(
            MenuItem(
                menu_id=menu.id,
                recipe_id=placement["recipe_id"],
                position=position,
                day_index=placement["day_index"],
                meal_slot=_slot_for_menu(menu_row, component_key),
                component_key=component_key,
                is_alternate=component_key == "alternate",
                source_type="seed",
            )
        )


def _replace_historical_menu_items(session, menu: HistoricalMenu, menu_row: dict) -> None:
    existing_items = session.scalars(
        select(HistoricalMenuItem).where(HistoricalMenuItem.historical_menu_id == menu.id)
    ).all()
    for item in existing_items:
        session.delete(item)

    for position, placement in enumerate(menu_row["items"], start=1):
        component_key = str(placement["component_key"])
        session.add(
            HistoricalMenuItem(
                historical_menu_id=menu.id,
                recipe_id=int(placement["recipe_id"]),
                position=position,
                day_index=int(placement["day_index"]),
                meal_slot=f"{menu_row['meal_type'].lower()}_{component_key}",
                component_key=component_key,
                is_alternate=component_key == "alternate",
                source_type="sample",
            )
        )


def _sync_historical_menu_metadata(menu: HistoricalMenu, menu_row: dict) -> None:
    """Keep seeded samples internally consistent after seed definitions change."""
    for field in [
        "name",
        "service_date",
        "program_type",
        "meal_type",
        "menu_coverage",
        "diet_type",
        "menu_duration_type",
        "meal_served_format",
        "menu_tags",
        "cycle",
        "days_per_week",
        "contracts",
        "sample_category",
        "passes_nutrition",
    ]:
        setattr(menu, field, menu_row[field])


def _sync_saved_menu_metadata(menu: Menu, menu_row: dict) -> None:
    for field in [
        "contract_name",
        "program_type",
        "meal_type",
        "menu_coverage",
        "diet_type",
        "menu_format",
        "menu_duration_type",
        "meal_served_format",
        "menu_tags",
        "cycle",
        "cycle_start_date",
        "cycle_end_date",
        "contracts",
        "completed_weeks",
        "submitted_programs",
        "status",
        "status_date",
        "submitted_to",
        "submitted_to_nyc_aging_on",
        "nutrition_advisor",
        "created_by",
        "service_date",
        "start_date",
        "end_date",
        "days_per_week",
        "cycle_week",
        "notes",
        "returned_comments",
        "approval_notes",
        "is_favorite",
    ]:
        setattr(menu, field, menu_row[field])


def seed_database() -> None:
    Base.metadata.create_all(bind=engine)
    session = SessionLocal()
    try:
        if session.scalar(select(HomeUpdate.id).limit(1)) is None:
            session.add_all([HomeUpdate(**row) for row in HOME_UPDATES])
        if session.scalar(select(RecipeHomeCategorySetting.category_key).limit(1)) is None:
            session.add_all([
                RecipeHomeCategorySetting(category_key=key, is_visible=is_visible)
                for key, is_visible in RECIPE_HOME_CATEGORIES
            ])
        if session.scalar(select(Recipe.recipe_id).limit(1)) is None:
            session.add_all([Recipe(**row) for row in RECIPES])
        else:
            for row in RECIPES:
                recipe = session.get(Recipe, row["recipe_id"])
                if recipe is None:
                    session.add(Recipe(**row))
                    continue
                for field in LEGACY_RECIPE_FIELDS:
                    value = getattr(recipe, field, None)
                    should_refresh_seed_ingredients = (
                        field == "ingredients"
                        and row["recipe_id"] <= 40
                        and not any(str(line).lstrip()[:1].isdigit() for line in (value or []))
                    )
                    if field in {"is_favorite", "vitamin_c_mg", "calcium_mg"} or value in (None, [], "") or should_refresh_seed_ingredients:
                        setattr(recipe, field, row[field])
        session.flush()
        _seed_recipe_attachments(session)
        if session.scalar(select(NutrientThreshold.id).limit(1)) is None:
            session.add_all([NutrientThreshold(**row) for row in THRESHOLDS])
        if session.scalar(select(HistoricalMenu.id).limit(1)) is None:
            for idx, menu in enumerate(HISTORICAL_MENUS, start=1):
                hist = HistoricalMenu(
                    name=menu["name"],
                    service_date=menu["service_date"],
                    program_type=menu["program_type"],
                    meal_type=menu["meal_type"],
                    menu_coverage=menu["menu_coverage"],
                    diet_type=menu["diet_type"],
                    menu_duration_type=menu["menu_duration_type"],
                    meal_served_format=menu["meal_served_format"],
                    menu_tags=menu["menu_tags"],
                    cycle=menu["cycle"],
                    days_per_week=menu["days_per_week"],
                    contracts=menu["contracts"],
                    sample_category=menu["sample_category"],
                    passes_nutrition=menu["passes_nutrition"],
                    notes="Seeded historical menu",
                )
                session.add(hist)
                session.flush()
                _replace_historical_menu_items(session, hist, menu)
        else:
            for index, row in enumerate(HISTORICAL_MENUS, start=1):
                hist = session.get(HistoricalMenu, index)
                if hist is None:
                    continue
                _sync_historical_menu_metadata(hist, row)
                _replace_historical_menu_items(session, hist, row)
        if session.scalar(select(Menu.id).limit(1)) is None:
            for menu_row in SAVED_MENUS:
                menu = Menu(
                    name=menu_row["name"],
                    contract_name=menu_row["contract_name"],
                    program_type=menu_row["program_type"],
                    meal_type=menu_row["meal_type"],
                    menu_coverage=menu_row["menu_coverage"],
                    diet_type=menu_row["diet_type"],
                    menu_format=menu_row["menu_format"],
                    menu_duration_type=menu_row["menu_duration_type"],
                    meal_served_format=menu_row["meal_served_format"],
                    menu_tags=menu_row["menu_tags"],
                    cycle=menu_row["cycle"],
                    cycle_start_date=menu_row["cycle_start_date"],
                    cycle_end_date=menu_row["cycle_end_date"],
                    contracts=menu_row["contracts"],
                    completed_weeks=menu_row["completed_weeks"],
                    submitted_programs=menu_row["submitted_programs"],
                    status=menu_row["status"],
                    status_date=menu_row["status_date"],
                    submitted_to=menu_row["submitted_to"],
                    submitted_to_nyc_aging_on=menu_row["submitted_to_nyc_aging_on"],
                    nutrition_advisor=menu_row["nutrition_advisor"],
                    created_by=menu_row["created_by"],
                    service_date=menu_row["service_date"],
                    start_date=menu_row["start_date"],
                    end_date=menu_row["end_date"],
                    days_per_week=menu_row["days_per_week"],
                    cycle_week=menu_row["cycle_week"],
                    notes=menu_row["notes"],
                    returned_comments=menu_row["returned_comments"],
                    approval_notes=menu_row["approval_notes"],
                    is_favorite=menu_row["is_favorite"],
                )
                session.add(menu)
                session.flush()
                _replace_saved_menu_items(session, menu, menu_row)
                for comment_row in menu_row["comments"]:
                    session.add(
                        MenuComment(
                            menu_id=menu.id,
                            created_at=comment_row["created_at"],
                            action=comment_row["action"],
                            author=comment_row["author"],
                            role=comment_row["role"],
                            body=comment_row["body"],
                            visibility=comment_row["visibility"],
                            badge_class=comment_row["badge_class"],
                            is_user_comment=True,
                            target_type=comment_row.get("target_type", "menu"),
                            target_label=comment_row.get("target_label", "General menu summary"),
                            day_index=comment_row.get("day_index"),
                            meal_slot=comment_row.get("meal_slot"),
                            component_key=comment_row.get("component_key"),
                            recipe_id=comment_row.get("recipe_id"),
                            nutrient_key=comment_row.get("nutrient_key"),
                            review_status=comment_row.get("review_status", "open"),
                        )
                    )
        elif session.scalar(select(MenuComment.id).limit(1)) is None:
            for menu_row in SAVED_MENUS:
                menu = session.scalar(select(Menu).where(Menu.name == menu_row["name"]))
                if menu is None:
                    continue
                _sync_saved_menu_metadata(menu, menu_row)
                _replace_saved_menu_items(session, menu, menu_row)
                for comment_row in menu_row["comments"]:
                    session.add(
                        MenuComment(
                            menu_id=menu.id,
                            created_at=comment_row["created_at"],
                            action=comment_row["action"],
                            author=comment_row["author"],
                            role=comment_row["role"],
                            body=comment_row["body"],
                            visibility=comment_row["visibility"],
                            badge_class=comment_row["badge_class"],
                            is_user_comment=True,
                            target_type=comment_row.get("target_type", "menu"),
                            target_label=comment_row.get("target_label", "General menu summary"),
                            day_index=comment_row.get("day_index"),
                            meal_slot=comment_row.get("meal_slot"),
                            component_key=comment_row.get("component_key"),
                            recipe_id=comment_row.get("recipe_id"),
                            nutrient_key=comment_row.get("nutrient_key"),
                            review_status=comment_row.get("review_status", "open"),
                        )
                    )
        else:
            for menu_row in SAVED_MENUS:
                menu = session.scalar(select(Menu).where(Menu.name == menu_row["name"]))
                if menu is None:
                    continue
                _sync_saved_menu_metadata(menu, menu_row)
                _replace_saved_menu_items(session, menu, menu_row)
                for comment_row in menu_row["comments"]:
                    existing_comment = session.scalar(
                        select(MenuComment).where(
                            MenuComment.menu_id == menu.id,
                            MenuComment.action == comment_row["action"],
                            MenuComment.body == comment_row["body"],
                        )
                    )
                    if existing_comment is None:
                        continue
                    existing_comment.target_type = comment_row.get("target_type", "menu")
                    existing_comment.target_label = comment_row.get("target_label", "General menu summary")
                    existing_comment.day_index = comment_row.get("day_index")
                    existing_comment.meal_slot = comment_row.get("meal_slot")
                    existing_comment.component_key = comment_row.get("component_key")
                    existing_comment.recipe_id = comment_row.get("recipe_id")
                    existing_comment.nutrient_key = comment_row.get("nutrient_key")
                    existing_comment.review_status = comment_row.get("review_status", "open")
        session.commit()
    finally:
        session.close()


if __name__ == "__main__":
    seed_database()
