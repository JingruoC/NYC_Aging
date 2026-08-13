from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import Recipe


HEADER_FILL = PatternFill("solid", fgColor="C75300")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(row) -> None:
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _add_key_value_sheet(workbook: Workbook, title: str, rows: list[tuple[str, object]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(["Field", "Value"])
    _style_header(sheet[1])
    for label, value in rows:
        sheet.append([label, value if value is not None else ""])
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 70


def _workbook_bytes(workbook: Workbook) -> bytes:
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def build_recipe_submission_workbook(recipe: Recipe) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_key_value_sheet(
        workbook,
        "Recipe submission",
        [
            ("Recipe ID", recipe.recipe_id),
            ("Recipe name", recipe.recipe_name),
            ("Meal type", recipe.meal_type),
            ("Category", recipe.category),
            ("Contributor", recipe.contributed_by),
            ("Created / uploaded", recipe.created_on),
            ("Serving size", recipe.serving_size),
            ("Yield / servings", recipe.yield_servings),
            ("Public", "Yes" if recipe.is_public else "No"),
            ("Approved", "Yes" if recipe.is_approved else "No"),
            ("Tags", ", ".join(recipe.tags or [])),
            ("Submission notes", recipe.scale_note),
        ],
    )

    ingredients = workbook.create_sheet("Ingredients")
    ingredients.append(["Line", "Ingredient"])
    _style_header(ingredients[1])
    for index, ingredient in enumerate(recipe.ingredients or [], start=1):
        ingredients.append([index, ingredient])
    ingredients.column_dimensions["A"].width = 10
    ingredients.column_dimensions["B"].width = 90

    directions = workbook.create_sheet("Directions")
    directions.append(["Step", "Direction"])
    _style_header(directions[1])
    for index, direction in enumerate(recipe.instructions or [], start=1):
        directions.append([index, direction])
    directions.column_dimensions["A"].width = 10
    directions.column_dimensions["B"].width = 100
    return _workbook_bytes(workbook)


def build_nutrition_workbook(recipe: Recipe) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _add_key_value_sheet(
        workbook,
        "Nutrition analysis",
        [
            ("Recipe ID", recipe.recipe_id),
            ("Recipe name", recipe.recipe_name),
            ("Serving size", recipe.serving_size),
            ("Calories", recipe.calories),
            ("Total fat (g)", recipe.fat_g),
            ("Saturated fat (g)", recipe.saturated_fat_g),
            ("Trans fat (g)", recipe.trans_fat_g),
            ("Cholesterol (mg)", recipe.cholesterol_mg),
            ("Sodium (mg)", recipe.sodium_mg),
            ("Total carbohydrate (g)", recipe.carbohydrates_g),
            ("Dietary fiber (g)", recipe.fiber_g),
            ("Total sugars (g)", recipe.total_sugars_g),
            ("Added sugars (g)", recipe.added_sugars_g),
            ("Protein (g)", recipe.protein_g),
            ("Vitamin D (mcg)", recipe.vitamin_d_mcg),
            ("Calcium (mg)", recipe.calcium_mg),
            ("Iron (mg)", recipe.iron_mg),
            ("Potassium (mg)", recipe.potassium_mg),
            ("Vitamin C (mg)", recipe.vitamin_c_mg),
        ],
    )
    return _workbook_bytes(workbook)
